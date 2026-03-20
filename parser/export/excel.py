from pathlib import Path
from threading import Lock
from datetime import datetime

from openpyxl import Workbook, load_workbook
from loguru import logger
from tzlocal import get_localzone

from parser.export.base import ResultStorage
from models import Item

class ExcelStorage(ResultStorage):
    """
    Сохранение результатов парсинга в XLSX
    """
    name = "excel"
    headers = [
        "Категория",
        "Название",
        "Характеристики объявления",
        "Прайс лист",
        "Цена",
        "URL",
        "Описание",
        "Дата публикации",
        "Имя продавца",
        "URL продавца",
        "Характеристики продавца",
        "Опыт работы",
        "Дата регистрации",
        "Количество всех объявлений",
        "Количество открытых объявлений",
        "Количество закрытых объяалений",
        "Тип продавца",
        "Адрес",
        "Адрес пользователя",
        "Координаты",
        "Изображения",
        "Количество изображений",
        "Видео",
        "Количество видео",
        "Поднято",
        "Просмотры (всего)",
        "Просмотры (на время сбора)",
        "Общая оценка",
        "Отзывы",
        "Количество отзывов",
        "Дата парсинга"
    ]

    def __init__(self, file_path: Path):
        self.file_path = file_path

        # создаём директорию
        self.file_path.parent.mkdir(parents=True, exist_ok=True)

        # lock для потокобезопасной записи
        self._lock = Lock()

        # создаём файл, если его нет
        if not self.file_path.exists():
            self._create_file()

    def _create_file(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(self.headers)
        workbook.save(self.file_path)

    @staticmethod
    def _get_ad_time(ad: Item):
        return (
            datetime
            .fromtimestamp(ad.sortTimeStamp / 1000, tz=get_localzone())
            .replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
        )

    @staticmethod
    def _get_item_coords(ad: Item) -> str:
        if ad.coords and "lat" in ad.coords and "lng" in ad.coords:
            return f"{ad.coords['lat']};{ad.coords['lng']}"
        return ""

    @staticmethod
    def _get_item_address_user(ad: Item) -> str:
        if ad.coords and "address_user" in ad.coords:
            return ad.coords["address_user"]
        return ""

    @staticmethod
    def _get_largest_image_url(img) -> str:
        try:
            best_key = max(
                img.root.keys(),
                key=lambda k: int(k.split("x")[0]) * int(k.split("x")[1])
            )
            return str(img.root[best_key])
        except Exception as err:
            logger.error(f"При определении лучшего изображения ошибка: {err}")
            return ""

    @staticmethod
    def excel_safe(value):
        # Formula Injection fix
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    def save(self, ads: list[Item]) -> None:
        if not ads:
            return

        with self._lock:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active

            for ad in ads:
                images_urls = [
                    self._get_largest_image_url(img)
                    for img in ad.images
                ]

                row = [
                    self.excel_safe(ad.category.specification or ""),
                    self.excel_safe(ad.title),
                    self.excel_safe(ad.additional_info or ""),
                    self.excel_safe(ad.price_list),
                    ad.priceDetailed.value,
                    self.excel_safe(f"https://www.avito.ru/{ad.urlPath}"),
                    self.excel_safe(ad.description),
                    self._get_ad_time(ad),
                    self.excel_safe(ad.seller.name if ad.seller else ""),
                    self.excel_safe(f"https://www.avito.ru/{ad.seller.url}" if ad.seller else ""),
                    self.excel_safe(ad.seller.characteristics if ad.seller else ""),
                    self.excel_safe(ad.seller.experience if ad.seller else ""),
                    self.excel_safe(ad.seller.registration_date if ad.seller else ""),
                    self.excel_safe(ad.seller.completed_ad_count + ad.seller.active_ad_count if ad.seller else ""),
                    self.excel_safe(ad.seller.active_ad_count if ad.seller else ""),
                    self.excel_safe(ad.seller.completed_ad_count if ad.seller else ""),
                    self.excel_safe(ad.seller.type if ad.seller else ""),
                    self.excel_safe(ad.location.name if ad.location else ""),
                    self.excel_safe(self._get_item_address_user(ad)),
                    self.excel_safe(self._get_item_coords(ad)),
                    self.excel_safe("\n".join(images_urls)),
                    self.excel_safe(str(len(images_urls))),
                    self.excel_safe("\n".join(ad.videos)),
                    self.excel_safe(str(len(ad.videos))),
                    "Да" if ad.isPromotion else "Нет",
                    ad.total_views or "",
                    ad.today_views or "",
                    self.excel_safe(ad.score or ""),
                    self.excel_safe(ad.reviews or ""),
                    ad.count_reviews or 0,
                    self.excel_safe(datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ]

                sheet.append(row)

                workbook.save(self.file_path)



